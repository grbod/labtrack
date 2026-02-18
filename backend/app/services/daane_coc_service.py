"""Service for generating Daane Labs Chain of Custody forms."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from zoneinfo import ZoneInfo
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional
import re
from xml.sax.saxutils import escape

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError

from app.models import (
    Lot,
    LotProduct,
    Product,
    ProductTestSpecification,
    LabTestType,
    RetestRequest,
    RetestItem,
    DaaneTestMapping,
    DaaneCOCDailyCounter,
    User,
)
from app.utils.logger import logger


@dataclass(frozen=True)
class DaaneMethodEntry:
    full: str
    name: str
    method: str
    turnaround: str
    normalized_name: str
    normalized_method: str


@dataclass
class DaaneTestRow:
    test_name: str
    test_method: Optional[str]
    specification: Optional[str]
    unit: Optional[str]
    daane_method: Optional[str]


class DaaneCOCService:
    """Generate Daane Labs COC XLSX files and manage test mappings."""

    MATCH_NAME_METHOD = "name_method"
    MATCH_NAME_ONLY = "name_only"
    MATCH_UNMAPPED = "unmapped"

    _method_list_cache: Optional[List[DaaneMethodEntry]] = None

    # Hardcoded company info per requirements
    COMPANY_NAME = "Body Nutrition"
    COMPANY_PHONE = "727-310-1916"
    COMPANY_ADDRESS = "2950 47th Ave. N, St. Petersburg, FL 33714"
    COMPANY_EMAIL = "Qa@bodynutrition.com, tvillegas@bodynutrition.com & greg@bodynutrition.com"
    DAANE_LAB_NAME = "Daane Labs"
    DAANE_LAB_ADDRESS = "4795 Enterprise Ave., Naples, FL 34104"
    DAANE_LAB_CONTACT = "239-227-4735 | login@daanelabs.com"
    DAANE_FORM_VERSION = "1.07F_Rev_08"

    def __init__(self) -> None:
        self.template_path = (
            Path(__file__).parent.parent.parent / "templates" / "daane_coc_v8.xlsx"
        )

    def ensure_mappings(self, db: Session) -> None:
        """Ensure mappings exist for all lab test types."""
        method_list = self._get_method_list()
        lab_tests = db.query(LabTestType).all()
        existing = {
            mapping.lab_test_type_id: mapping
            for mapping in db.query(DaaneTestMapping).all()
        }

        updated = False
        for test in lab_tests:
            mapping = existing.get(test.id)
            if mapping:
                if mapping.daane_method is None:
                    daane_method, match_type, reason = self._suggest_daane_method(
                        test.test_name, test.test_method, method_list
                    )
                    mapping.daane_method = daane_method
                    mapping.match_type = match_type
                    mapping.match_reason = reason
                    updated = True
                continue

            daane_method, match_type, reason = self._suggest_daane_method(
                test.test_name, test.test_method, method_list
            )
            db.add(
                DaaneTestMapping(
                    lab_test_type_id=test.id,
                    daane_method=daane_method,
                    match_type=match_type,
                    match_reason=reason,
                )
            )
            updated = True

        if updated:
            db.commit()

    def rebuild_mappings(self, db: Session) -> None:
        """Rebuild mappings for all lab test types."""
        self._method_list_cache = None
        method_list = self._get_method_list()
        lab_tests = db.query(LabTestType).all()
        existing = {
            mapping.lab_test_type_id: mapping
            for mapping in db.query(DaaneTestMapping).all()
        }

        for test in lab_tests:
            daane_method, match_type, reason = self._suggest_daane_method(
                test.test_name, test.test_method, method_list
            )
            mapping = existing.get(test.id)
            if mapping:
                mapping.daane_method = daane_method
                mapping.match_type = match_type
                mapping.match_reason = reason
            else:
                db.add(
                    DaaneTestMapping(
                        lab_test_type_id=test.id,
                        daane_method=daane_method,
                        match_type=match_type,
                        match_reason=reason,
                    )
                )

        db.commit()

    def list_mappings(self, db: Session):
        """Return all mappings with lab test type info."""
        self.ensure_mappings(db)
        mappings = (
            db.query(DaaneTestMapping)
            .join(LabTestType, DaaneTestMapping.lab_test_type_id == LabTestType.id)
            .order_by(LabTestType.test_name.asc())
            .all()
        )
        return mappings

    def generate_coc_for_lot(self, db: Session, lot_id: int, user: User) -> tuple[bytes, int]:
        """Generate a Daane COC XLSX for a lot."""
        lot = (
            db.query(Lot)
            .options(
                joinedload(Lot.lot_products)
                .joinedload(LotProduct.product)
                .joinedload(Product.test_specifications)
                .joinedload(ProductTestSpecification.lab_test_type)
            )
            .filter(Lot.id == lot_id)
            .first()
        )

        if not lot:
            raise ValueError(f"Lot with ID {lot_id} not found")

        po_number = self._po_number_for_lot(db, lot)
        sample_id = lot.reference_number
        lot_number = lot.lot_number

        tests = self._build_tests_for_lot(db, lot)
        serving_size_note = self._serving_size_note(lot.products)

        return self._render_coc(
            po_number=po_number,
            authorizer_name=self._authorizer_name(user),
            authorizer_signature=self._authorizer_signature(user),
            sample_id=sample_id,
            lot_number=lot_number,
            method_suitability="None",
            special_instructions=serving_size_note,
            tests=tests,
        ), len(tests)

    def generate_coc_pdf_for_lot(
        self,
        db: Session,
        lot_id: int,
        user: User,
        selected_lab_test_type_ids: Optional[List[int]] = None,
        special_instructions: Optional[str] = None,
    ) -> tuple[bytes, int]:
        """Generate a Daane COC PDF for a lot."""
        lot = (
            db.query(Lot)
            .options(
                joinedload(Lot.lot_products)
                .joinedload(LotProduct.product)
                .joinedload(Product.test_specifications)
                .joinedload(ProductTestSpecification.lab_test_type)
            )
            .filter(Lot.id == lot_id)
            .first()
        )

        if not lot:
            raise ValueError(f"Lot with ID {lot_id} not found")

        po_number = self._po_number_for_lot(db, lot)
        sample_id = lot.reference_number
        lot_number = lot.lot_number

        selected_ids_set: Optional[set[int]] = None
        if selected_lab_test_type_ids is not None:
            selected_ids_set = {int(test_id) for test_id in selected_lab_test_type_ids}
            if not selected_ids_set:
                raise ValueError("At least one test must be selected to generate a Daane COC PDF")

            available_ids = self._lot_lab_test_type_ids(lot)
            invalid_ids = sorted(selected_ids_set - available_ids)
            if invalid_ids:
                raise ValueError(
                    f"Selected tests are not available for this lot: {', '.join(str(i) for i in invalid_ids)}"
                )

        tests = self._build_tests_for_lot(db, lot, selected_lab_test_type_ids=selected_ids_set)
        instructions = special_instructions if special_instructions is not None else self._serving_size_note(lot.products)

        return self._render_coc_pdf(
            po_number=po_number,
            authorizer_name=self._authorizer_name(user),
            authorizer_signature=self._authorizer_signature(user),
            sample_id=sample_id,
            lot_number=lot_number,
            method_suitability="None",
            special_instructions=instructions,
            tests=tests,
        ), len(tests)

    def generate_coc_for_retest(
        self, db: Session, request_id: int
    ) -> tuple[bytes, int]:
        """Generate a Daane COC XLSX for a retest request."""
        retest_request = (
            db.query(RetestRequest)
            .options(
                joinedload(RetestRequest.lot)
                .joinedload(Lot.lot_products)
                .joinedload(LotProduct.product),
                joinedload(RetestRequest.requested_by),
                joinedload(RetestRequest.items).joinedload(RetestItem.test_result),
            )
            .filter(RetestRequest.id == request_id)
            .first()
        )

        if not retest_request:
            raise ValueError(f"Retest request with ID {request_id} not found")

        po_number = self._po_number_for_retest(db, retest_request)

        sample_id = retest_request.reference_number
        lot_number = retest_request.lot.lot_number if retest_request.lot else ""
        serving_size_note = (
            self._serving_size_note(retest_request.lot.products)
            if retest_request.lot
            else ""
        )
        requested_by = retest_request.requested_by
        authorizer_name = self._authorizer_name(requested_by) if requested_by else ""
        authorizer_signature = self._authorizer_signature(requested_by) if requested_by else ""

        tests = self._build_tests_for_retest(db, retest_request)

        return self._render_coc(
            po_number=po_number,
            authorizer_name=authorizer_name,
            authorizer_signature=authorizer_signature,
            sample_id=sample_id,
            lot_number=lot_number,
            method_suitability="None",
            special_instructions=serving_size_note,
            tests=tests,
        ), len(tests)

    def generate_coc_pdf_for_retest(
        self, db: Session, request_id: int, special_instructions: Optional[str] = None,
    ) -> tuple[bytes, int]:
        """Generate a Daane COC PDF for a retest request."""
        retest_request = (
            db.query(RetestRequest)
            .options(
                joinedload(RetestRequest.lot)
                .joinedload(Lot.lot_products)
                .joinedload(LotProduct.product),
                joinedload(RetestRequest.requested_by),
                joinedload(RetestRequest.items).joinedload(RetestItem.test_result),
            )
            .filter(RetestRequest.id == request_id)
            .first()
        )

        if not retest_request:
            raise ValueError(f"Retest request with ID {request_id} not found")

        po_number = self._po_number_for_retest(db, retest_request)

        sample_id = retest_request.reference_number
        lot_number = retest_request.lot.lot_number if retest_request.lot else ""
        serving_size_note = (
            self._serving_size_note(retest_request.lot.products)
            if retest_request.lot
            else ""
        )
        requested_by = retest_request.requested_by
        authorizer_name = self._authorizer_name(requested_by) if requested_by else ""
        authorizer_signature = self._authorizer_signature(requested_by) if requested_by else ""

        tests = self._build_tests_for_retest(db, retest_request)
        instructions = special_instructions if special_instructions is not None else serving_size_note

        return self._render_coc_pdf(
            po_number=po_number,
            authorizer_name=authorizer_name,
            authorizer_signature=authorizer_signature,
            sample_id=sample_id,
            lot_number=lot_number,
            method_suitability="None",
            special_instructions=instructions,
            tests=tests,
        ), len(tests)

    def _render_coc(
        self,
        *,
        po_number: str,
        authorizer_name: str,
        authorizer_signature: str,
        sample_id: str,
        lot_number: str,
        method_suitability: str,
        special_instructions: str,
        tests: List[DaaneTestRow],
    ) -> bytes:
        """Render the COC XLSX with provided values."""
        if not self.template_path.exists():
            raise FileNotFoundError(f"Daane COC template not found at {self.template_path}")

        wb = openpyxl.load_workbook(self.template_path)
        ws = wb["Chain of Custody"]

        # Customer information
        ws["D3"] = self.COMPANY_NAME
        ws["M3"] = self.COMPANY_EMAIL
        ws["C4"] = self.COMPANY_PHONE
        ws["C5"] = self.COMPANY_ADDRESS
        ws["C6"] = po_number
        ws["O5"] = authorizer_name
        ws["O6"] = authorizer_signature

        # Sample row (use first row only)
        ws["D9"] = sample_id
        ws["P9"] = lot_number
        ws["S9"] = method_suitability

        # Special instructions
        if special_instructions:
            ws["D19"] = special_instructions

        # Test rows
        start_row = 22
        max_rows = 12
        for idx, test in enumerate(tests[:max_rows]):
            row = start_row + idx
            ws[f"A{row}"] = test.daane_method or ""
            ws[f"O{row}"] = test.specification or ""
            ws[f"T{row}"] = test.unit or ""

        # Warn if tests exceed available rows
        if len(tests) > max_rows:
            logger.warning(
                f"Daane COC: {len(tests)} tests provided, only {max_rows} rendered"
            )

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _render_coc_pdf(
        self,
        *,
        po_number: str,
        authorizer_name: str,
        authorizer_signature: str,
        sample_id: str,
        lot_number: str,
        method_suitability: str,
        special_instructions: str,
        tests: List[DaaneTestRow],
    ) -> bytes:
        """Render a PDF version of the COC."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )
        styles = getSampleStyleSheet()
        def _para(text: Optional[str], style: ParagraphStyle) -> Paragraph:
            safe = escape(text or "")
            safe = safe.replace("\n", "<br/>")
            return Paragraph(safe, style)

        title_style = ParagraphStyle(
            "DaaneTitle",
            parent=styles["Heading1"],
            alignment=TA_CENTER,
            fontSize=16,
            spaceAfter=6,
        )
        section_style = ParagraphStyle(
            "DaaneSection",
            parent=styles["Heading3"],
            alignment=TA_LEFT,
            fontSize=11,
            spaceBefore=6,
            spaceAfter=4,
        )
        header_style = ParagraphStyle(
            "DaaneHeader",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
        )
        header_right_style = ParagraphStyle(
            "DaaneHeaderRight",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            alignment=TA_RIGHT,
        )
        value_style = ParagraphStyle(
            "DaaneValue",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
        )
        label_style = ParagraphStyle(
            "DaaneLabel",
            parent=styles["BodyText"],
            fontSize=9,
            leading=11,
            fontName="Helvetica-Bold",
        )
        email_style = ParagraphStyle(
            "DaaneEmail",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
        )
        small_style = ParagraphStyle(
            "DaaneSmall",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
        )

        story: List = []
        header_left = Paragraph(
            f"<b>{escape(self.DAANE_LAB_NAME)}</b><br/>{escape(self.DAANE_LAB_ADDRESS)}<br/>{escape(self.DAANE_LAB_CONTACT)}",
            header_style,
        )
        header_right = _para(self.DAANE_FORM_VERSION, header_right_style)
        header_table = Table(
            [[header_left, header_right]],
            colWidths=[5.6 * inch, 1.9 * inch],
        )
        header_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        story.append(header_table)
        story.append(Spacer(1, 0.08 * inch))
        story.append(Paragraph("Chain of Custody", title_style))

        # Customer information table
        email_lines = self.COMPANY_EMAIL.split(",")
        email_lines = [line.strip() for line in email_lines if line.strip()]
        email_text = "<br/>".join(escape(line) for line in email_lines)
        address_text = self.COMPANY_ADDRESS.replace(", ", "\n", 1)

        customer_data = [
            ["Customer Information", "", "", ""],
            [
                _para("Company Name:", label_style),
                _para(self.COMPANY_NAME, value_style),
                _para("Email:", label_style),
                Paragraph(email_text, email_style),
            ],
            [
                _para("Phone:", label_style),
                _para(self.COMPANY_PHONE, value_style),
                _para("Authorizer Name:", label_style),
                _para(authorizer_name, value_style),
            ],
            [
                _para("Address:", label_style),
                _para(address_text, value_style),
                _para("Authorizer Signature:", label_style),
                _para(authorizer_signature, value_style),
            ],
            [_para("PO #:", label_style), _para(po_number, value_style), "", ""],
        ]
        customer_table = Table(
            customer_data,
            colWidths=[1.35 * inch, 2.4 * inch, 1.05 * inch, 2.7 * inch],
        )
        customer_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (-1, 0)),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#94a3b8")),
                    ("INNERGRID", (0, 1), (-1, -1), 0.3, colors.HexColor("#cbd5f5")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(customer_table)

        story.append(Spacer(1, 0.12 * inch))

        # Sample table
        sample_header = ["Sample ID / Description", "Lot Number", "Method Suit. #"]
        sample_data = [
            sample_header,
            [
                _para(sample_id, value_style),
                _para(lot_number, value_style),
                _para(method_suitability, value_style),
            ],
        ]
        sample_table = Table(sample_data, colWidths=[4.1 * inch, 1.8 * inch, 1.6 * inch])
        sample_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("LINEABOVE", (0, 0), (-1, 0), 0.6, colors.HexColor("#94a3b8")),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.4, colors.HexColor("#cbd5f5")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94a3b8")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5f5")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(sample_table)

        story.append(Spacer(1, 0.12 * inch))

        # Special instructions
        story.append(Paragraph("Special Instructions", section_style))
        instructions_text = special_instructions or ""
        instructions_table = Table([[_para(instructions_text, value_style)]], colWidths=[7.5 * inch])
        instructions_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94a3b8")),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("MINHEIGHT", (0, 0), (-1, -1), 0.5 * inch),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(instructions_table)

        story.append(Spacer(1, 0.12 * inch))

        # Test selection table
        test_header_label = self._template_cell_value("Chain of Custody", "A21")
        if test_header_label:
            story.append(Paragraph(test_header_label.split("\n")[0], section_style))
        else:
            story.append(Paragraph("Test Selection", section_style))

        test_data = [["Test Selection / Method", "Specifications", "Units"]]
        for test in tests[:12]:
            test_data.append(
                [
                    _para(test.daane_method or "", value_style),
                    _para(test.specification or "", value_style),
                    _para(test.unit or "", value_style),
                ]
            )

        test_table = Table(
            test_data,
            colWidths=[4.4 * inch, 1.8 * inch, 1.3 * inch],
        )
        test_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("LINEABOVE", (0, 0), (-1, 0), 0.6, colors.HexColor("#94a3b8")),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.4, colors.HexColor("#cbd5f5")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#94a3b8")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5f5")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(test_table)

        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    def _build_tests_for_lot(
        self,
        db: Session,
        lot: Lot,
        selected_lab_test_type_ids: Optional[set[int]] = None,
    ) -> List[DaaneTestRow]:
        self.ensure_mappings(db)
        mappings = self._mapping_dict(db)
        tests: List[DaaneTestRow] = []

        specs_by_lab_test_id: Dict[int, ProductTestSpecification] = {}
        for product in lot.products:
            for spec in product.test_specifications:
                if (
                    selected_lab_test_type_ids is not None
                    and spec.lab_test_type_id not in selected_lab_test_type_ids
                ):
                    continue
                if spec.lab_test_type_id not in specs_by_lab_test_id:
                    specs_by_lab_test_id[spec.lab_test_type_id] = spec

        for spec in specs_by_lab_test_id.values():
            lab_test = spec.lab_test_type
            daane_method = self._resolve_daane_method(
                mappings,
                lab_test.id if lab_test else None,
                lab_test.test_name if lab_test else None,
                lab_test.test_method if lab_test else None,
            )
            tests.append(
                DaaneTestRow(
                    test_name=lab_test.test_name if lab_test else "",
                    test_method=lab_test.test_method if lab_test else None,
                    specification=spec.specification,
                    unit=lab_test.default_unit if lab_test else None,
                    daane_method=daane_method,
                )
            )

        return tests

    def _lot_lab_test_type_ids(self, lot: Lot) -> set[int]:
        ids: set[int] = set()
        for product in lot.products:
            for spec in product.test_specifications:
                ids.add(spec.lab_test_type_id)
        return ids

    def _build_tests_for_retest(
        self, db: Session, retest_request: RetestRequest
    ) -> List[DaaneTestRow]:
        self.ensure_mappings(db)
        mappings = self._mapping_dict(db)
        tests: List[DaaneTestRow] = []

        for item in retest_request.items:
            test_result = item.test_result
            if not test_result:
                continue

            lab_test = (
                db.query(LabTestType)
                .filter(LabTestType.test_name.ilike(test_result.test_type))
                .first()
            )
            daane_method = self._resolve_daane_method(
                mappings,
                lab_test.id if lab_test else None,
                lab_test.test_name if lab_test else test_result.test_type,
                lab_test.test_method if lab_test else test_result.method,
            )

            tests.append(
                DaaneTestRow(
                    test_name=test_result.test_type,
                    test_method=test_result.method,
                    specification=test_result.specification,
                    unit=test_result.unit,
                    daane_method=daane_method,
                )
            )

        return tests

    def _mapping_dict(self, db: Session) -> Dict[int, DaaneTestMapping]:
        return {
            mapping.lab_test_type_id: mapping
            for mapping in db.query(DaaneTestMapping).all()
        }

    def _resolve_daane_method(
        self,
        mappings: Dict[int, DaaneTestMapping],
        lab_test_type_id: Optional[int],
        test_name: Optional[str],
        test_method: Optional[str],
    ) -> Optional[str]:
        if lab_test_type_id and lab_test_type_id in mappings:
            mapping = mappings[lab_test_type_id]
            if mapping.daane_method:
                return mapping.daane_method

        method_list = self._get_method_list()
        daane_method, _, _ = self._suggest_daane_method(test_name, test_method, method_list)
        return daane_method

    def _serving_size_note(self, products: List[Product]) -> str:
        serving_sizes = {p.serving_size for p in products if p.serving_size}
        if not serving_sizes:
            return "(NO SERVING SIZE DETERMINED)"
        sizes = ", ".join(sorted(serving_sizes))
        return f"Serving Size = {sizes}"

    def _authorizer_name(self, user: Optional[User]) -> str:
        if not user:
            return ""
        return user.full_name or user.username

    def _authorizer_signature(self, user: Optional[User]) -> str:
        if not user:
            return ""
        name = user.full_name or user.username
        return f"/s/ {name}"

    def _po_number_for_lot(self, db: Session, lot: Lot) -> str:
        if lot.daane_po_number:
            return lot.daane_po_number
        po_number = self._format_po_number(self._today_eastern(), self._next_daily_sequence(db))
        lot.daane_po_number = po_number
        db.commit()
        return po_number

    def _po_number_for_retest(self, db: Session, retest_request: RetestRequest) -> str:
        if retest_request.daane_po_number:
            return retest_request.daane_po_number
        po_number = self._format_po_number(self._today_eastern(), self._next_daily_sequence(db))
        retest_request.daane_po_number = po_number
        db.commit()
        return po_number

    def generate_po_number(self, db: Session) -> str:
        """Generate a new PO number using the daily counter (Eastern time)."""
        return self._format_po_number(self._today_eastern(), self._next_daily_sequence(db))

    def _next_daily_sequence(self, db: Session) -> int:
        base_date = self._today_eastern()
        counter = (
            db.query(DaaneCOCDailyCounter)
            .filter(DaaneCOCDailyCounter.counter_date == base_date)
            .first()
        )
        if counter:
            counter.last_sequence += 1
            db.commit()
            return counter.last_sequence

        try:
            counter = DaaneCOCDailyCounter(counter_date=base_date, last_sequence=1)
            db.add(counter)
            db.commit()
            return 1
        except IntegrityError:
            db.rollback()
            counter = (
                db.query(DaaneCOCDailyCounter)
                .filter(DaaneCOCDailyCounter.counter_date == base_date)
                .first()
            )
            if not counter:
                raise
            counter.last_sequence += 1
            db.commit()
            return counter.last_sequence

    def _today_eastern(self) -> date:
        return datetime.now(ZoneInfo("America/New_York")).date()

    def _format_po_number(self, base_date: date, sequence: int) -> str:
        julian_day = base_date.strftime("%j")
        year_two = base_date.strftime("%y")
        return f"PO-{year_two}{julian_day}-{sequence:03d}"

    def _template_cell_value(self, sheet_name: str, cell: str) -> Optional[str]:
        if not self.template_path.exists():
            return None
        wb = openpyxl.load_workbook(self.template_path, data_only=True)
        ws = wb[sheet_name]
        value = ws[cell].value
        if isinstance(value, str):
            return value.strip()
        return None

    def _get_method_list(self) -> List[DaaneMethodEntry]:
        if self._method_list_cache is not None:
            return self._method_list_cache

        if not self.template_path.exists():
            logger.error(f"Daane COC template not found: {self.template_path}")
            self._method_list_cache = []
            return self._method_list_cache

        wb = openpyxl.load_workbook(self.template_path, data_only=True)
        ws = wb["Method List"]

        entries: List[DaaneMethodEntry] = []
        seen = set()
        for row in ws.iter_rows(min_row=1, max_col=3):
            cell_value = row[2].value
            if not isinstance(cell_value, str):
                continue
            full = cell_value.strip()
            if not full or full in seen:
                continue
            seen.add(full)

            parts = [part.strip() for part in full.split(" / ")]
            name = parts[0] if parts else full
            method = parts[1] if len(parts) > 1 else ""
            turnaround = parts[2] if len(parts) > 2 else ""
            entries.append(
                DaaneMethodEntry(
                    full=full,
                    name=name,
                    method=method,
                    turnaround=turnaround,
                    normalized_name=self._normalize_text(name),
                    normalized_method=self._normalize_text(method),
                )
            )

        self._method_list_cache = entries
        return entries

    def _suggest_daane_method(
        self,
        test_name: Optional[str],
        test_method: Optional[str],
        method_list: List[DaaneMethodEntry],
    ) -> Tuple[Optional[str], str, Optional[str]]:
        if not test_name:
            return None, self.MATCH_UNMAPPED, "missing test name"

        normalized_test = self._normalize_text(test_name)
        if not normalized_test:
            return None, self.MATCH_UNMAPPED, "empty normalized test name"

        candidates: List[Tuple[float, DaaneMethodEntry]] = []
        for entry in method_list:
            score = self._name_similarity(normalized_test, entry.normalized_name)
            if score >= 0.6:
                candidates.append((score, entry))

        if not candidates:
            return None, self.MATCH_UNMAPPED, "no name match"

        candidates.sort(key=lambda item: item[0], reverse=True)

        if test_method:
            normalized_method = self._normalize_text(test_method)
            for score, entry in candidates:
                if normalized_method and (
                    normalized_method in entry.normalized_method
                    or SequenceMatcher(
                        None, normalized_method, entry.normalized_method
                    ).ratio()
                    >= 0.7
                ):
                    return entry.full, self.MATCH_NAME_METHOD, f"name+method match ({entry.method})"

        best_entry = candidates[0][1]
        return best_entry.full, self.MATCH_NAME_ONLY, "name match only"

    def _name_similarity(self, normalized_test: str, normalized_name: str) -> float:
        if not normalized_test or not normalized_name:
            return 0.0
        if normalized_test in normalized_name or normalized_name in normalized_test:
            return 1.0
        return SequenceMatcher(None, normalized_test, normalized_name).ratio()

    def _normalize_text(self, value: Optional[str]) -> str:
        if not value:
            return ""
        text = value.lower()
        text = text.replace("&", "and")
        text = re.sub(r"[^a-z0-9]+", "", text)
        return text


daane_coc_service = DaaneCOCService()
